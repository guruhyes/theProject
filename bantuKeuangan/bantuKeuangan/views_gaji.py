from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import load_workbook
def index(request):
    context={
        'judul':'Gaji'
    }
    return render(request,'gaji.html',context)
def generateExcel(request,form):
    if(form=="form01"):
        wb = load_workbook("./need/form01.xlsx")
        ws = wb['ket_gaji_pot']
        ws['B10'] = "Gaji Bulan Maret 2222"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=ff.xlsx'
        wb.save(response)
        return response